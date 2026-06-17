"""M5 助手：fs_access 授权/越界 + 文件问答 + 办公文档 + draft。"""
from __future__ import annotations

import pytest

from core.harness.fs_access import FolderAccess
from core.tools.assistant import (
    DocxWriteInput,
    DocxWriteTool,
    DraftInput,
    DraftTool,
    PptxBuildInput,
    PptxBuildTool,
    ReadDirInput,
    ReadDirTool,
    SearchFilesInput,
    SearchFilesTool,
    SummarizeFilesInput,
    SummarizeFilesTool,
    XlsxWriteInput,
    XlsxWriteTool,
)
from tests.helpers import make_ctx


def _noop(_chunk):
    return None


# ---------- fs_access ----------
def test_fs_authorize_resolve(tmp_path):
    fs = FolderAccess()
    rec = fs.authorize(str(tmp_path))
    assert rec["mode"] == "read_write"
    assert str(fs.resolve(str(tmp_path / "a.txt"))).startswith(str(tmp_path.resolve()))


def test_fs_escape_blocked(tmp_path):
    fs = FolderAccess()
    fs.authorize(str(tmp_path / "docs"))
    with pytest.raises(PermissionError):
        fs.resolve(str(tmp_path / "secret.txt"))
    with pytest.raises(PermissionError):
        fs.resolve(str(tmp_path / "docs" / ".." / "secret.txt"))  # ../ 越界


def test_fs_read_only_blocks_write(tmp_path):
    fs = FolderAccess()
    fs.authorize(str(tmp_path), mode="read_only")
    fs.resolve(str(tmp_path / "a"), need_write=False)  # 读 ok
    with pytest.raises(PermissionError):
        fs.resolve(str(tmp_path / "a"), need_write=True)


def test_fs_persistence(tmp_path):
    sp = str(tmp_path / "folders.json")
    FolderAccess(sp).authorize(str(tmp_path / "d"))
    fs2 = FolderAccess(sp)
    assert len(fs2.list()) == 1
    fs2.revoke(fs2.list()[0]["id"])
    assert FolderAccess(sp).list() == []


# ---------- 文件问答 ----------
async def test_read_dir(tmp_path):
    (tmp_path / "a.txt").write_text("x")
    (tmp_path / "sub").mkdir()
    fs = FolderAccess()
    fs.authorize(str(tmp_path))
    res = await ReadDirTool(fs).call(ReadDirInput(path=str(tmp_path)), make_ctx(tmp_path), _noop)
    assert "a.txt" in res.data and "sub" in res.data


async def test_read_dir_unauthorized(tmp_path):
    vr = await ReadDirTool(FolderAccess()).validate_input(
        ReadDirInput(path=str(tmp_path)), make_ctx(tmp_path)
    )
    assert not vr.ok and "授权" in vr.message


async def test_search_files(tmp_path):
    (tmp_path / "a.txt").write_text("hello world\nfoo")
    (tmp_path / "b.txt").write_text("nothing")
    fs = FolderAccess()
    fs.authorize(str(tmp_path))
    res = await SearchFilesTool(fs).call(
        SearchFilesInput(query="hello", path=str(tmp_path)), make_ctx(tmp_path), _noop
    )
    assert "a.txt" in res.data and "hello" in res.data


async def test_summarize_files(tmp_path):
    (tmp_path / "a.txt").write_text("内容A")
    fs = FolderAccess()
    fs.authorize(str(tmp_path))
    res = await SummarizeFilesTool(fs).call(
        SummarizeFilesInput(paths=[str(tmp_path / "a.txt")]), make_ctx(tmp_path), _noop
    )
    assert "内容A" in res.data


async def test_draft():
    res = await DraftTool().call(DraftInput(kind="邮件", content="你好"), None, _noop)
    assert "邮件" in res.data and "你好" in res.data


# ---------- 办公文档（真生成 + 读回）----------
async def test_xlsx_write(tmp_path):
    from openpyxl import load_workbook

    fs = FolderAccess()
    fs.authorize(str(tmp_path))
    out = str(tmp_path / "t.xlsx")
    await XlsxWriteTool(fs).call(
        XlsxWriteInput(path=out, rows=[["a", "b"], [1, 2]]), make_ctx(tmp_path), _noop
    )
    ws = load_workbook(out).active
    assert ws["A1"].value == "a" and ws["B2"].value == 2


async def test_docx_write(tmp_path):
    from docx import Document

    fs = FolderAccess()
    fs.authorize(str(tmp_path))
    out = str(tmp_path / "t.docx")
    await DocxWriteTool(fs).call(
        DocxWriteInput(path=out, title="标题", paragraphs=["p1", "p2"]), make_ctx(tmp_path), _noop
    )
    texts = [p.text for p in Document(out).paragraphs]
    assert "标题" in texts and "p1" in texts


async def test_pptx_build(tmp_path):
    from pptx import Presentation

    fs = FolderAccess()
    fs.authorize(str(tmp_path))
    out = str(tmp_path / "t.pptx")
    await PptxBuildTool(fs).call(
        PptxBuildInput(path=out, slides=[{"title": "T", "bullets": ["x", "y"]}]),
        make_ctx(tmp_path), _noop,
    )
    assert len(Presentation(out).slides) == 1


async def test_office_write_unauthorized_blocked(tmp_path):
    vr = await XlsxWriteTool(FolderAccess()).validate_input(
        XlsxWriteInput(path=str(tmp_path / "t.xlsx"), rows=[]), make_ctx(tmp_path)
    )
    assert not vr.ok
